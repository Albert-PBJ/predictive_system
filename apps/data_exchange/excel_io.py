"""Lectura y escritura de los archivos .xlsx (openpyxl).

Tres operaciones: **leer** la hoja de una operación a filas crudas por clave de columna,
construir la **plantilla** en blanco (con instrucciones, listas desplegables y hojas de
referencia con los catálogos de apoyo) y construir el **export** de datos existentes.
"""

from __future__ import annotations

import unicodedata
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .handlers import Handler, reference_sheet_rows

_HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_REQ_FILL = PatternFill("solid", fgColor="EFF4FF")


def _norm_header(text: str) -> str:
    s = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode("ascii")
    return " ".join(s.lower().split())


class SheetError(Exception):
    """Problema estructural del archivo (hoja o columnas faltantes)."""


# --------------------------------------------------------------------------- #
# Lectura
# --------------------------------------------------------------------------- #
def read_entity_sheet(file, handler: Handler) -> list[dict]:
    """Lee la hoja de la operación y devuelve filas crudas ``{"_row": n, key: valor}``.

    Empareja las cabeceras del archivo con las columnas del handler (ignorando mayúsculas
    y acentos). Lanza ``SheetError`` si falta una columna obligatoria o el archivo no es
    un .xlsx legible.
    """
    try:
        wb = load_workbook(file, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 — archivo corrupto o no-xlsx
        raise SheetError(f"No se pudo leer el archivo Excel (.xlsx). Detalle: {exc}") from exc

    ws = wb[handler.label] if handler.label in wb.sheetnames else wb[wb.sheetnames[0]]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise SheetError(f"La hoja «{ws.title}» está vacía.")

    index_by_norm = {}
    for idx, val in enumerate(header_row):
        if val is not None and str(val).strip():
            index_by_norm[_norm_header(val)] = idx

    col_index, missing = {}, []
    for c in handler.columns:
        idx = index_by_norm.get(_norm_header(c.header))
        if idx is None:
            if c.required:
                missing.append(c.header)
        else:
            col_index[c.key] = idx
    if missing:
        raise SheetError(
            f"Faltan columnas obligatorias en la hoja «{ws.title}»: {', '.join(missing)}. "
            "Descarga la plantilla y respeta los nombres de las columnas."
        )

    rows = []
    for r_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        raw = {"_row": r_i}
        for key, idx in col_index.items():
            raw[key] = row[idx] if idx < len(row) else None
        rows.append(raw)
    wb.close()
    return rows


# --------------------------------------------------------------------------- #
# Escritura — plantilla y export
# --------------------------------------------------------------------------- #
def _cell_value(value):
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    return value


def _write_headers(ws, handler: Handler):
    for c_i, col in enumerate(handler.columns, start=1):
        cell = ws.cell(row=1, column=c_i, value=col.header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        letter = get_column_letter(c_i)
        ws.column_dimensions[letter].width = max(14, min(40, len(col.header) + 6))
        note = col.help + (" (obligatorio)" if col.required else "")
        if note.strip():
            cell.comment = Comment(note, "Sistema")
    ws.freeze_panes = "A2"


def _apply_dropdowns(ws, handler: Handler):
    for c_i, col in enumerate(handler.columns, start=1):
        if not col.dropdown:
            continue
        letter = get_column_letter(c_i)
        dv = DataValidation(
            type="list", formula1='"%s"' % ",".join(col.dropdown), allow_blank=True,
        )
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}1000")


def _write_reference_sheet(wb, kind: str):
    rows = reference_sheet_rows(kind)
    if not rows:
        return
    ws = wb.create_sheet(kind.capitalize())
    for r_i, row in enumerate(rows, start=1):
        for c_i, val in enumerate(row, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=_cell_value(val))
            if r_i == 1:
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
    for c_i in range(1, len(rows[0]) + 1):
        ws.column_dimensions[get_column_letter(c_i)].width = 24
    ws.freeze_panes = "A2"


def _write_instructions(ws, handler: Handler):
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 70
    ws.cell(row=1, column=1, value=f"Plantilla de importación — {handler.label}").font = _TITLE_FONT
    r = 3
    for line in handler.instructions:
        c = ws.cell(row=r, column=1, value="•")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.cell(row=r, column=1, value=f"•  {line}").alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 15 * (len(line) // 90 + 1))
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Columna").font = _HEADER_FONT
    ws.cell(row=r, column=2, value="Obligatoria").font = _HEADER_FONT
    ws.cell(row=r, column=3, value="Notas").font = _HEADER_FONT
    for c_i in (1, 2, 3):
        ws.cell(row=r, column=c_i).fill = _HEADER_FILL
    r += 1
    for col in handler.columns:
        ws.cell(row=r, column=1, value=col.header)
        ws.cell(row=r, column=2, value="Sí" if col.required else "No")
        notes = col.help
        if col.dropdown:
            notes = (notes + "  " if notes else "") + "Valores: " + ", ".join(col.dropdown)
        ws.cell(row=r, column=3, value=notes).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1


def build_template_workbook(handler: Handler) -> BytesIO:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Instrucciones"
    _write_instructions(ws0, handler)

    ws = wb.create_sheet(handler.label)
    _write_headers(ws, handler)
    _apply_dropdowns(ws, handler)

    for kind in handler.reference_sheets:
        _write_reference_sheet(wb, kind)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def build_export_workbook(handler: Handler, rows: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = handler.label
    _write_headers(ws, handler)
    for r_i, row in enumerate(rows, start=2):
        for c_i, col in enumerate(handler.columns, start=1):
            ws.cell(row=r_i, column=c_i, value=_cell_value(row.get(col.key)))
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
