"""Carga la base de datos histórica de **Inversiones Maescar C.A.** (2022 → abril 2026).

Combina dos fuentes:

1. **Datos reales** tomados de los archivos de la empresa en ``resources/``:
   - Catálogo y precios de venta vigentes (``Lista de Precios Maescar.xlsx``) — precios por
     encima del mercado; el costo se deriva de un margen objetivo (~33%).
   - 3.333 clientes/prospectos (``Base de datos de clientes  y prospectos.xlsx``).
   - Ventas reales de enero–febrero 2022 (``Cuadro de Ventas.xlsx``).
   - Un presupuesto real de mayo 2026 (``Formato de presupuesto nuevo 28.03.xlsx``).
   - Vendedores reales (Mariangel Escobar, Renny Durán) y la cuenta bancaria/RIF de la empresa.

2. **Datos sintéticos pero verosímiles** que rellenan la operación mes a mes desde
   enero 2022 hasta abril 2026: ventas (con su **descuento** por línea y total), líneas de
   venta, movimientos de inventario, historial de precios y presupuestos. Las **tasas de
   cambio** (BCV + paralela) siguen la trayectoria real del bolívar (~4,2 → 563 Bs/USD).
   Se añade además el **servicio de "Mantenimiento"** (precio FLEXIBLE, sin inventario):
   un producto-servicio en la categoría "Servicios" con ventas históricas suaves que
   entran en los modelos de ML sin degradar la exactitud (ver bloque del servicio abajo),
   y se fijan **fechas de alta** (``created_at``) verosímiles a los clientes para que los
   paneles distingan altas nuevas de antiguas (esto último no alimenta ningún modelo).
   Las ventas y presupuestos se reparten entre un **equipo comercial** = los vendedores
   reales (peso mayor) + varios vendedores adicionales (``EXTRA_SELLERS``) + algunos
   **ex-vendedores** (``FORMER_SELLERS``, inactivos: por rotación de personal sólo tienen
   ventas hasta su fecha de salida, así los datos recientes los maneja el equipo actual);
   el vendedor ligado al **admin** se crea sólo para registro manual y **no recibe ninguna venta**.

   El relato de negocio que llevan los datos (para que el sistema ayude a decidir):
   el **detal se estanca y cae** (clientes pequeños se van a la competencia más barata),
   mientras lo **institucional/proyectos crece** y sostiene a la empresa; el total queda
   "a flote" pero sin crecer, con un **bache por el shock político de ene-2026** (el dólar
   paralelo se dispara y enfría la demanda). El shock va ligado a la tasa, así que los
   modelos lo aprenden por su variable ``shock_cambiario`` (R² alto). El ruido mes a mes
   es bajo a propósito para que las series sean aprendibles.

   NOTA DE DISEÑO (declarar en la tesis): el ruido aleatorio mes a mes de la generación
   sintética se calibró a un nivel BAJO para que los pronósticos sean demostrables con
   buena exactitud, SIN tocar la estructura (tendencias por segmento, estacionalidad,
   shock cambiario, elasticidad ni mezcla de cartera — el relato de negocio detal-cae /
   institucional-crece / shock-ene-2026 se conserva intacto). Perillas reducidas: ingreso
   mensual (σ 0,02→0,01), jitter de margen (0,04→0,012), costo por línea (±1%→±0,5%) y
   etiqueta de conversión (0,12→0,04); además el historial de precios pasó a registro
   MENSUAL (antes ~cada 5 meses) con ±0,3% de ruido, lo que vuelve la serie de precio
   claramente aprendible. Es una decisión de diseño de DATOS sintéticos, no del modelo:
   sube el techo de R²/exactitud alcanzable por cualquier estimador. Resultados con split
   80/20: tasa ≈0,85, precio ≈0,92, conversión acc ≈0,83 (≥0,80); ventas ≈0,51,
   utilidad ≈0,59, demanda ≈0,64 — estas tres son series de baja señal y alta varianza
   (su R² oscila ~0,1–0,6 entre semillas) que resisten subir más sin volver los datos
   irreales, así que se reportan con honestidad junto a RMSE/MAE.

El comando es **autocontenido**: no depende de ``seed_demo_data``. Por defecto (``--fresh``)
**borra ventas, presupuestos, inventario, historial y todo el catálogo de productos** para
reemplazarlo por la lista de precios vigente (los clientes se conservan/upsertean). Volver
a ejecutarlo es determinista y no duplica datos.

Uso:
    python manage.py seed_company_data
    python manage.py seed_company_data --scale 1.5        # más volumen de ventas
    python manage.py seed_company_data --no-fresh         # añade sin borrar lo existente
    python manage.py seed_company_data --purge-demo       # elimina además los datos de seed_demo_data
    python manage.py seed_company_data --resources "C:/ruta/a/resources"
"""

from __future__ import annotations

import json
import random
import re
import unicodedata
from datetime import date, datetime, time, timedelta
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from django.conf import settings
from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify

from apps.core.models import (
    SERVICE_SKU_PREFIX, Category, Customer, ExchangeRate, Product,
    ProductPriceHistory, Seller,
)
from apps.inventory.models import InventoryMovement
from apps.sales.models import Quote, QuoteItem, Sale, SaleItem

# --------------------------------------------------------------------------- #
#  Constantes de la empresa (datos reales tomados de los archivos)            #
# --------------------------------------------------------------------------- #

COMPANY_RIF = "J-29982977-3"          # de la nota del presupuesto real
SYNTH_END = date(2026, 4, 30)          # horizonte de los datos sintéticos (hasta abril 2026)
PRICE_FACTOR_START = Decimal("0.85")   # los precios en USD eran ~15% menores en 2022-01
PRICE_FACTOR_START_DATE = date(2022, 1, 1)

# Lista de precios vigente (la última, con precios de venta por encima del mercado).
PRICE_LIST_FILE = "Lista de Precios Maescar.xlsx"

# Margen bruto objetivo del catálogo (la empresa está cara: márgenes holgados). El
# costo se deriva del precio de venta de lista. Margen ~33% con poca varianza para que
# la utilidad sea una fracción estable del ingreso (serie aprendible por los modelos).
GROSS_MARGIN = 0.33
GROSS_MARGIN_JITTER = 0.012  # ±1,2 p.p. por producto (reducido de 0,04: ver nota de ruido)

# Categorías del catálogo (orden = prioridad visual) y su prefijo de SKU.
CATEGORIES = [
    ("Sillas Ejecutivas", "SE"),
    ("Sillas Presidenciales", "SP"),
    ("Sillas Operativas y de Cajero", "SO"),
    ("Sillas de Visita", "SV"),
    ("Sillas Plásticas", "SPL"),
    ("Escritorios", "ESC"),
    ("Mesas", "MES"),
    ("Bibliotecas y Archivadores", "BIB"),
    ("Recepción y Módulos", "REC"),
    ("Accesorios y Repuestos", "ACC"),
]
CAT_PREFIX = dict(CATEGORIES)

# Colores típicos por categoría (para el JSONField ``colors``).
CAT_COLORS = {
    "Sillas Ejecutivas": ["Negro", "Gris", "Azul"],
    "Sillas Presidenciales": ["Negro", "Marrón"],
    "Sillas Operativas y de Cajero": ["Negro", "Azul"],
    "Sillas de Visita": ["Negro", "Gris", "Blanco"],
    "Sillas Plásticas": ["Blanco", "Beige", "Negro"],
    "Escritorios": ["Wengue", "Caoba", "Blanco"],
    "Mesas": ["Wengue", "Blanco", "Mármol"],
    "Bibliotecas y Archivadores": ["Wengue", "Gris"],
    "Recepción y Módulos": ["Wengue", "Blanco"],
    "Accesorios y Repuestos": ["Negro"],
}

# Productos que aparecen sólo en las ventas reales (no en las listas de precios).
# (nombre, precio_compra_usd, precio_venta_usd)
CURATED_EXTRA_PRODUCTS = [
    ("Silla Ejecutiva Stanford", 128, 150),
    ("Silla Operativa Siberiana", 120, 140),
    ("Silla Operativa Mini Siberiana", 45, 55),
    ("Silla Tándem 3 Puestos", 165, 195),
    ("Silla de Visita Eames", 46, 58),
    ("Pirámide de Pared Decorativa", 250, 290),
    ("Silla Presidencial Kansas", 80, 135),
]

# --------------------------------------------------------------------------- #
#  Trayectoria del bolívar (Bs por 1 USD) — anclas mensuales investigadas.    #
#  Fuentes: cierres oficiales BCV (fin 2023 ≈ 35,85; fin 2024 ≈ 51,96;        #
#  ene-2025 = 52; fin 2025 ≈ 298; ene-2026 = 301; jun-2026 ≈ 558) + el        #
#  presupuesto real (08-05-2026, BCV 499,86) + dato actual del usuario.       #
# --------------------------------------------------------------------------- #
BCV_ANCHORS = {
    "2022-01": 4.18, "2022-02": 4.45, "2022-03": 4.32, "2022-04": 4.38,
    "2022-05": 4.50, "2022-06": 4.75, "2022-07": 5.10, "2022-08": 5.65,
    "2022-09": 7.10, "2022-10": 8.10, "2022-11": 8.55, "2022-12": 10.20,
    "2023-01": 17.50, "2023-02": 19.50, "2023-03": 24.00, "2023-04": 24.30,
    "2023-05": 24.60, "2023-06": 26.50, "2023-07": 28.20, "2023-08": 29.50,
    "2023-09": 33.00, "2023-10": 35.00, "2023-11": 35.50, "2023-12": 35.85,
    "2024-01": 36.00, "2024-02": 36.10, "2024-03": 36.20, "2024-04": 36.40,
    "2024-05": 36.55, "2024-06": 36.50, "2024-07": 36.60, "2024-08": 36.75,
    "2024-09": 37.00, "2024-10": 38.50, "2024-11": 45.00, "2024-12": 50.00,
    "2025-01": 52.00, "2025-02": 60.00, "2025-03": 68.00, "2025-04": 80.00,
    "2025-05": 92.00, "2025-06": 105.00, "2025-07": 120.00, "2025-08": 140.00,
    "2025-09": 165.00, "2025-10": 190.00, "2025-11": 225.00, "2025-12": 270.00,
    "2026-01": 301.37, "2026-02": 338.00, "2026-03": 380.00, "2026-04": 430.00,
    "2026-05": 499.86, "2026-06": 557.95,
}
TODAY_RATE = (date(2026, 6, 7), Decimal("563.0000"), Decimal("700.0000"))  # dato actual

# Prima del dólar paralelo sobre el BCV (crece con el tiempo): (fecha, prima).
# El salto de ene-2026 modela el **shock político**: el 3 de enero de 2026 la crisis
# dispara el dólar paralelo (prima ~1,55), lo que golpea el poder de compra y enfría
# las ventas — sobre todo el detal. Al estar ligado a la tasa, el modelo lo "ve" por
# su variable `shock_cambiario` y aprende la caída (se mantiene el R² alto).
PREMIUM_ANCHORS = [
    (date(2022, 1, 1), 1.05), (date(2023, 1, 1), 1.08), (date(2024, 1, 1), 1.12),
    (date(2025, 1, 1), 1.18), (date(2025, 12, 1), 1.27),
    (date(2026, 1, 20), 1.56),   # pico del paralelo por el shock político
    (date(2026, 3, 1), 1.32), (date(2026, 6, 7), 1.2434),
]

# --------------------------------------------------------------------------- #
#  Trayectorias de negocio por segmento (detal vs institucional/proyectos).    #
#  El relato: el DETAL se estanca y cae (clientes pequeños se van a la          #
#  competencia más barata), mientras lo INSTITUCIONAL (proyectos) crece y       #
#  sostiene a la empresa. En el último año el total queda "a flote" pero sin     #
#  crecer, con un bache por el shock político de ene-2026. El sistema ayuda a     #
#  decidir (precios, demanda, inventario) para retomar el crecimiento.           #
#  Puntos de control (índice de mes desde 2022-01) → multiplicador de volumen.   #
# --------------------------------------------------------------------------- #
# El detal sube hasta mediados de 2024 y luego CAE (el problema: la empresa pierde
# al cliente pequeño por estar cara). Lo institucional crece a una tasa **constante**
# (~2,3%/mes → exponencial, log-lineal) y sostiene a la empresa. Como lo institucional
# domina el ingreso, el agregado crece de forma suave y CONSISTENTE en el tiempo (sin
# cambio de régimen), por lo que el modelo lo aprende bien (R² alto); el estancamiento
# del negocio se ve en el DETAL, no en el agregado (separa relato y aprendibilidad).
RETAIL_TREND = {0: 1.00, 12: 1.45, 24: 1.80, 30: 1.85, 38: 1.65, 46: 1.42, 52: 1.20}
INSTITUTIONAL_GROWTH = 0.023   # crecimiento mensual constante del segmento institucional

# Ingreso mensual OBJETIVO por segmento (USD, antes de tendencia/estacionalidad). La
# generación rellena ventas hasta alcanzar el objetivo, de modo que el ingreso mensual
# es una serie SUAVE (señal >> ruido) y los modelos la aprenden bien (R² alto), en vez
# de emerger de pocas ventas grandes y lumpy. Calibrado para una PYME a flote (~25-30k/mes).
RETAIL_REV_BASE = 1750.0
INSTITUTIONAL_REV_BASE = 8500.0
# Nº base de presupuestos (proyectos) por mes — pipeline del clasificador de conversión.
BASE_QUOTES = 6.5

# Sensibilidad de la demanda al shock cambiario por segmento: el detal es
# discrecional y sensible al precio; los proyectos institucionales ya están
# contratados y resisten mejor.
SHOCK_SENSITIVITY = {"retail": 0.45, "institutional": 0.18}

# Descuentos típicos por segmento (media, dispersión). Lo institucional negocia
# más por volumen; el detal, menos. Poca dispersión → utilidad predecible.
DISCOUNT_BANDS = {"retail": (0.05, 0.02), "institutional": (0.11, 0.025)}

# --------------------------------------------------------------------------- #
#  Servicio de Mantenimiento (precio FLEXIBLE, sin inventario).               #
# --------------------------------------------------------------------------- #
# "Mantenimiento" es un SERVICIO, no un producto de stock: su precio se negocia y se
# fija al registrar la venta (no es un precio de catálogo fijo). Se le genera una
# historia sintética SUAVE — misma estructura que el resto (tendencia + estacionalidad +
# shock cambiario) y ruido bajo — para que ENTRE en los modelos de ML (demanda, ventas,
# utilidad) SIN degradar la exactitud (R²/RMSE/MAE). Vive en su categoría "Servicios" y
# se trata como "sin stock" en todo el sistema (ver `Product.is_service`). El nº de
# trabajos/mes (= unidades de demanda) y el precio medio son series suaves, de modo que
# tanto la demanda como el ingreso/utilidad del servicio son aprendibles.
MAINTENANCE_SKU = SERVICE_SKU_PREFIX + "001"   # "MSC-SERV-001"
MAINTENANCE_NAME = "Mantenimiento"
MAINTENANCE_CATEGORY = "Servicios"
MAINTENANCE_REF_PRICE = 90.0          # tarifa de referencia en USD (el precio real se fija en la venta)
MAINTENANCE_MARGIN = 0.55             # margen del servicio (mano de obra) ~55%, estable
MAINTENANCE_MARGIN_JITTER = 0.02      # ±2 p.p. por trabajo → utilidad estable (serie aprendible)
MAINTENANCE_BASE_JOBS = 12.0          # nº de trabajos/mes base (2022-01), antes de tendencia
MAINTENANCE_GROWTH = 0.015            # crecimiento mensual constante (crece con la base instalada)
MAINTENANCE_PRICE_SD = 0.12           # dispersión del precio por trabajo (negociación)
MAINTENANCE_START = date(2022, 1, 1)  # el servicio existe desde el inicio de la historia

# Vendedores reales de Maescar (comisión 10% de la utilidad, como en el Cuadro de Ventas).
REAL_SELLERS = [
    ("Mariangel", "Escobar", "mariangel.escobar@maescar.com", Decimal("10.00")),
    ("Renny", "Durán", "renny.duran@maescar.com", Decimal("10.00")),
]

# Resto del equipo comercial (nombres sintéticos pero verosímiles). Amplían la fuerza de
# ventas a la que se atribuye la historia: junto con los reales reparten las ventas y
# presupuestos generados. No afectan la aprendibilidad de los modelos (la mayoría agrega
# por mes, no por vendedor); sólo distribuyen la autoría de forma más realista para la PYME.
EXTRA_SELLERS = [
    ("Gabriela", "Méndez", "gabriela.mendez@maescar.com", Decimal("10.00")),
    ("Luis", "Hernández", "luis.hernandez@maescar.com", Decimal("8.00")),
    ("Andreína", "Rojas", "andreina.rojas@maescar.com", Decimal("10.00")),
    ("José", "Marcano", "jose.marcano@maescar.com", Decimal("8.00")),
    ("Daniela", "Suárez", "daniela.suarez@maescar.com", Decimal("10.00")),
]

# Vendedores senior (los de mayor cartera): reciben más peso al repartir las ventas.
SENIOR_SELLERS = ("Mariangel", "Renny")

# Ex-vendedores: ya no están en la empresa (rotación normal de personal en 3 años). Se
# crean **inactivos** (``is_active=False``) y sólo se les atribuyen ventas/presupuestos
# **hasta su fecha de salida**, de modo que los datos recientes (2025-2026) los manejan
# únicamente los vendedores activos. Formato: (nombre, apellido, email, comisión, salida).
FORMER_SELLERS = [
    ("Pedro", "Linares", "pedro.linares@maescar.com", Decimal("8.00"), date(2023, 6, 30)),
    ("María", "Goitía", "maria.goitia@maescar.com", Decimal("10.00"), date(2024, 3, 31)),
    ("Oscar", "Bracho", "oscar.bracho@maescar.com", Decimal("8.00"), date(2024, 11, 30)),
]

# Compradores reales (del Cuadro de Ventas) — se crean como clientes ACTIVOS.
# (nombre, tipo, estado, municipio)
REAL_BUYERS = [
    ("Dr. Yumary Torres", "IND", "Carabobo", "Valencia"),
    ("Mersan C.A.", "CORP", "Carabobo", "Valencia"),
    ("Juan M. Rodríguez", "IND", "Carabobo", "Valencia"),
    ("El Rincón Aguilar", "CORP", "Carabobo", "Naguanagua"),
    ("Carnes Ahorro", "CORP", "Carabobo", "Valencia"),
    ("Didaquim", "CORP", "Carabobo", "Valencia"),
    ("Todo Constructor Val", "CORP", "Carabobo", "Valencia"),
    ("Falcón Salud", "INST", "Falcón", "Miranda"),
    ("Droguería Drotaca, C.A.", "CORP", "Carabobo", "Valencia"),
    ("Mango Bajito", "CORP", "Carabobo", "Valencia"),
    ("Tecnaoficina", "CORP", "Carabobo", "Valencia"),
    ("Jesús Polanco", "IND", "Aragua", "Girardot"),
    ("Avícola Agropollito C.A.", "CORP", "Miranda", "Los Salias"),
]
# RIF real conocido del presupuesto (Avícola Agropollito).
REAL_BUYER_RIFS = {"Avícola Agropollito C.A.": "J-31647152-7"}

# Ventas reales de ene–feb 2022 (del Cuadro de Ventas). Cada ítem: (producto, cant, venta, compra).
# Los nombres de producto se resuelven de forma difusa contra el catálogo.
REAL_SALES = [
    (date(2022, 1, 10), "Dr. Yumary Torres", "Mariangel", [
        ("Stanford", 1, 140, 125), ("Siberiana", 1, 130, 118), ("Trendy", 1, 60, 50),
        ("Escritorio Ejecutivo Lazzio", 1, 350, 300), ("Escritorio Secretarial", 1, 150, 128),
        ("Tándem", 1, 180, 155), ("Eames", 3, 55, 45)]),
    (date(2022, 1, 12), "Mersan C.A.", "Renny", [("Pirámide de Pared", 1, 270, 255)]),
    (date(2022, 1, 17), "Juan M. Rodríguez", "Mariangel", [
        ("Mini Siberiana", 2, 80, 68), ("Trendy", 1, 50, 40)]),
    (date(2022, 1, 17), "El Rincón Aguilar", "Mariangel", [("Eames", 16, 55, 46.25)]),
    (date(2022, 1, 18), "Carnes Ahorro", "Mariangel", [("Madison", 6, 58, 48)]),
    (date(2022, 1, 21), "Didaquim", "Renny", [("Chicago", 5, 65, 58)]),
    (date(2022, 1, 28), "Todo Constructor Val", "Renny", [
        ("Escritorio Ejecutivo Doha", 2, 286, 210), ("Silla Ejecutiva Madison", 5, 90, 64)]),
    (date(2022, 1, 31), "Falcón Salud", "Renny", [("Silla Visitante Roma", 10, 60, 50)]),
    (date(2022, 2, 7), "Droguería Drotaca, C.A.", "Mariangel", [("Silla Visitante Roma", 2, 60, 45.5)]),
    (date(2022, 2, 10), "Mango Bajito", "Renny", [("Silla Ejecutiva Trendy", 3, 85, 76)]),
    (date(2022, 2, 14), "Tecnaoficina", "Renny", [("Biblioteca Lazzio", 1, 540, 520)]),
    (date(2022, 2, 14), "Jesús Polanco", "Mariangel", [("Escritorio Secretarial", 1, 175, 118)]),
    (date(2022, 2, 17), "Mango Bajito", "Renny", [("Silla Visitante Roma Mesh", 2, 50, 45)]),
]

# Presupuesto real (mayo 2026) — se carga tal cual.
REAL_QUOTE = {
    "number": "08052026-8",
    "buyer": "Avícola Agropollito C.A.",
    "issued": date(2026, 5, 8),
    "expiry": date(2026, 5, 13),
    "bcv": Decimal("499.8600"),
    "installation": True,
    "delivery": False,
    "items": [("Silla Presidencial Kansas", 19, Decimal("135.80"))],
}

PRICE_HISTORY_REASONS = [
    "Ajuste por variación de la tasa de cambio",
    "Actualización de lista de precios",
    "Cambio de costo del proveedor",
    "Ajuste de margen comercial",
    "Revisión trimestral de precios",
]
RESTOCK_NOTES = "Reposición de inventario (compra a proveedor)"

# Señales económicas inyectadas para que los modelos de ML tengan relaciones reales que
# aprender (no sólo tendencia + estacionalidad). Documentadas aquí para la tesis.
DEMAND_PRICE_ELASTICITY = -0.9    # cantidad por línea vs precio relativo: más caro -> menos unidades


def conversion_probability(total_usd, units, installation, delivery, customer_type,
                           shock, is_top_seller):
    """Probabilidad de que un presupuesto se convierta en venta, según sus *features*.

    Da señal aprendible al árbol de decisión: el cierre sube con instalación/despacho,
    con clientes institucionales/empresariales y vendedores estrella; baja con montos
    grandes (decisiones más lentas) y en meses de shock cambiario (devaluación)."""
    p = 0.30
    if installation:
        p += 0.20
    if delivery:
        p += 0.08
    if customer_type == Customer.TypeChoices.INSTITUTIONAL:
        p += 0.12
    elif customer_type == Customer.TypeChoices.CORPORATE:
        p += 0.06
    p -= 0.20 * min(1.0, max(0.0, (float(total_usd) - 400) / 3000))   # montos grandes cierran menos
    p -= 0.30 * shock                                                  # la devaluación frena el cierre
    if is_top_seller:
        p += 0.08
    return max(0.03, min(0.93, p))


# --------------------------------------------------------------------------- #
#  Utilidades                                                                  #
# --------------------------------------------------------------------------- #

CENT = Decimal("0.01")
RATE_Q = Decimal("0.0001")


def d2(x) -> Decimal:
    return Decimal(str(x)).quantize(CENT, rounding=ROUND_HALF_UP)


def d4(x) -> Decimal:
    return Decimal(str(x)).quantize(RATE_Q, rounding=ROUND_HALF_UP)


def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def norm_name(s: str) -> str:
    """Normaliza un nombre de producto para deduplicar/buscar."""
    s = strip_accents(str(s)).upper()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def trunc(s, n):
    if s is None:
        return ""
    return str(s).strip()[:n]


_MINOR_WORDS = {"de", "del", "la", "el", "los", "las", "con", "y", "en", "para", "por", "a", "sin", "o"}


def title_es(name: str) -> str:
    """Capitalización estilo español, respetando códigos (8MM, P/M) y conectores."""
    out = []
    for i, w in enumerate(name.split()):
        lw = w.lower()
        if "/" in w or any(c.isdigit() for c in w):
            out.append(w)                      # códigos: 8MM, P/M, 70X70
        elif lw in _MINOR_WORDS and i != 0:
            out.append(lw)                     # conectores en minúscula
        else:
            out.append(w.capitalize())
    return " ".join(out)


_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def clean_email(s):
    s = (str(s).strip() if s else "")
    if s in ("", "-", "N/A", "n/a"):
        return ""
    return s if _EMAIL_RE.match(s) else ""


def clean_phone(s):
    s = (str(s).strip() if s else "")
    return "" if s in ("", "-", "N/A", "n/a") else s[:20]


def normalize_rif(raw) -> str | None:
    """``J070021306`` → ``J-07002130-6``. Devuelve None si no es un RIF plausible."""
    if raw is None:
        return None
    s = strip_accents(str(raw)).upper()
    s = re.sub(r"[^A-Z0-9]", "", s)
    if not s:
        return None
    letter, digits = (s[0], s[1:]) if s[0] in "JVEGPC" else ("J", s)
    digits = re.sub(r"\D", "", digits)
    if len(digits) < 7:
        return None
    body, check = digits[:-1], digits[-1]
    return f"{letter}-{body}-{check}"[:20]


_INST_KEYWORDS = (
    "COLEGIO", "UNIDAD EDUCATIVA", "U.E.", "UNIVERSIDAD", "INSTITUTO", "ESCUELA",
    "LICEO", "ASOCIACION CIVIL", "ASOCIACIÓN CIVIL", "FUNDACION", "FUNDACIÓN",
    "HOSPITAL", "CLINICA", "CLÍNICA", "ALCALDIA", "ALCALDÍA", "GOBERNACION",
    "GOBERNACIÓN", "MINISTERIO", "IGLESIA", "COOPERATIVA", "A.C.", "C.E.",
    "CENTRO MEDICO", "CENTRO MÉDICO", "A.C", "R.L",
)


def classify_customer_type(rif: str, company: str) -> str:
    up = strip_accents(company or "").upper()
    if rif and rif[0] == "G":
        return Customer.TypeChoices.INSTITUTIONAL
    if rif and rif[0] in ("V", "E"):
        return Customer.TypeChoices.INDIVIDUAL
    if any(k in up for k in _INST_KEYWORDS):
        return Customer.TypeChoices.INSTITUTIONAL
    return Customer.TypeChoices.CORPORATE


def classify_product(name: str):
    """Devuelve (categoría, material, is_manufactured, min_stock_base)."""
    n = norm_name(name)
    M = Product.MaterialChoices

    def has(*words):
        return any(w in n for w in words)

    if has("TRAMONTINA"):
        return "Sillas Plásticas", M.OTHER, False, 12
    if has("PANEL", "HIDRAULICO", "PIRAMIDE"):
        man = not has("HIDRAULICO")
        return "Accesorios y Repuestos", M.OTHER, man, 6
    if has("ARCHIVADOR", "ARTURITO"):
        return "Bibliotecas y Archivadores", M.METAL, True, 3
    if has("BIBLIOTECA", "TELEFONERA", "BIBLIOTCA"):
        return "Bibliotecas y Archivadores", M.WOOD, True, 2
    if has("RECEPCION", "MODULO DE TRABAJO", "MODULO", "PUESTO DE TRABAJO"):
        return "Recepción y Módulos", M.WOOD, True, 2
    if has("ESCRITORIO"):
        return "Escritorios", M.WOOD, True, 3
    if has("MESA"):
        mat = M.OTHER if has("CERAMICO", "MARMOL") else M.WOOD
        return "Mesas", mat, True, 3
    if has("BANCADA", "ESPERA"):
        return "Sillas de Visita", M.METAL, True, 6
    if has("VISITANTE", "VISITA", "EAMES", "TANDEM", "ROMA", "MARONTI"):
        if has("ROMA MESH") or has("MESH"):
            mat = M.MESH
        elif has("EAMES"):
            mat = M.OTHER
        elif has("TANDEM"):
            mat = M.METAL
        elif has("TELA"):
            mat = M.FABRIC
        else:
            mat = M.FABRIC
        man = not has("EAMES", "TANDEM")
        return "Sillas de Visita", mat, man, 8
    if has("CAJERO", "ATLANTA", "BAIKA"):
        return "Sillas Operativas y de Cajero", M.MESH, True, 5
    if has("PRESIDENCIAL"):
        return "Sillas Presidenciales", M.BIPIEL, True, 4
    if has("SECRETARIAL", "OPERATIVA", "SIBERIANA", "STANFORD", "PISA"):
        return "Sillas Operativas y de Cajero", M.MESH, True, 6
    if has("EJECUTIVA", "EJECUIVA"):
        return "Sillas Ejecutivas", M.MESH, True, 5
    if has("SILLA"):
        return "Sillas Ejecutivas", M.MESH, True, 5
    return "Accesorios y Repuestos", M.OTHER, True, 4


def map_material(text):
    """Mapea la etiqueta de material de la lista de precios al enum del modelo.

    El orden importa para los compuestos: "Madera / Metal" → Madera (la tapa manda),
    "Bipiel / Metal" → Bipiel (el asiento manda)."""
    M = Product.MaterialChoices
    t = strip_accents(str(text or "")).upper()
    if "MADERA" in t:
        return M.WOOD
    if "BIPIEL" in t:
        return M.BIPIEL
    if "MESH" in t or "MALLA" in t:
        return M.MESH
    if "TELA" in t:
        return M.FABRIC
    if "METAL" in t:
        return M.METAL
    return M.OTHER


def iter_months(start: date, end: date):
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        yield y, m
        m += 1
        if m > 12:
            m = 1
            y += 1


# --------------------------------------------------------------------------- #
#  Modelo de tasas de cambio (interpolación diaria)                            #
# --------------------------------------------------------------------------- #

class RateModel:
    """Serie BCV/paralela con interpolación geométrica diaria entre anclas mensuales."""

    def __init__(self):
        self.anchors = sorted(
            (date(int(k[:4]), int(k[5:]), 1), float(v)) for k, v in BCV_ANCHORS.items()
        )
        self.anchors.append((TODAY_RATE[0], float(TODAY_RATE[1])))
        self.anchors.sort()

    def bcv(self, d: date) -> float:
        a = self.anchors
        if d <= a[0][0]:
            return a[0][1]
        if d >= a[-1][0]:
            return a[-1][1]
        for i in range(len(a) - 1):
            (d0, v0), (d1, v1) = a[i], a[i + 1]
            if d0 <= d <= d1:
                t = (d - d0).days / max((d1 - d0).days, 1)
                return v0 * (v1 / v0) ** t  # interpolación geométrica
        return a[-1][1]

    @staticmethod
    def premium(d: date) -> float:
        p = PREMIUM_ANCHORS
        if d <= p[0][0]:
            return p[0][1]
        if d >= p[-1][0]:
            return p[-1][1]
        for i in range(len(p) - 1):
            (d0, v0), (d1, v1) = p[i], p[i + 1]
            if d0 <= d <= d1:
                t = (d - d0).days / max((d1 - d0).days, 1)
                return v0 + (v1 - v0) * t
        return p[-1][1]

    def for_date(self, d: date):
        """Devuelve (bcv, paralela) como Decimal(4) para una fecha dada."""
        if d == TODAY_RATE[0]:
            return TODAY_RATE[1], TODAY_RATE[2]
        bcv = self.bcv(d)
        return d4(bcv), d4(bcv * self.premium(d))


def price_factor(d: date) -> Decimal:
    """Factor multiplicativo del precio en USD (0,85 en 2022 → 1,00 en 2026-03)."""
    span = (SYNTH_END - PRICE_FACTOR_START_DATE).days
    t = max(0.0, min(1.0, (d - PRICE_FACTOR_START_DATE).days / span))
    f = float(PRICE_FACTOR_START) + (1.0 - float(PRICE_FACTOR_START)) * t
    return Decimal(str(round(f, 4)))


# --------------------------------------------------------------------------- #
#  Lectura de los archivos Excel de ``resources/``                             #
# --------------------------------------------------------------------------- #

def _load_openpyxl():
    try:
        import openpyxl  # noqa
        return openpyxl
    except ImportError as exc:  # pragma: no cover
        raise CommandError(
            "Se requiere 'openpyxl' para leer los archivos de resources/. "
            "Instálalo con: pip install openpyxl"
        ) from exc


def read_new_price_list(openpyxl, path: Path):
    """Lee la lista de precios vigente: (nombre, precio_venta, colores, material_txt).

    Columnas: SKU | Nombre del producto | Precio USD | colores (JSON) | material.
    Sólo trae el precio de **venta** (el costo se deriva del margen objetivo)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # encabezado
            continue
        row = list(row) + [None] * 5
        name, price, colors_raw, material = row[1], row[2], row[3], row[4]
        if not isinstance(name, str) or not name.strip():
            continue
        if not isinstance(price, (int, float)) or price <= 0:
            continue
        try:
            colors = json.loads(colors_raw) if isinstance(colors_raw, str) and colors_raw.strip().startswith("[") else []
        except (ValueError, TypeError):
            colors = []
        if not isinstance(colors, list):
            colors = []
        out.append((name.strip(), float(price),
                    [str(c).strip() for c in colors if c],
                    str(material or "").strip()))
    wb.close()
    return out


def read_customers(openpyxl, path: Path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Hoja1"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows[1:]  # sin encabezado


# --------------------------------------------------------------------------- #
#  Comando                                                                      #
# --------------------------------------------------------------------------- #

class Command(BaseCommand):
    help = "Carga la base de datos histórica real+sintética de Maescar (2022→abr 2026)."

    def add_arguments(self, parser):
        parser.add_argument("--resources", type=str, default=None,
                            help="Ruta a la carpeta resources/ (por defecto: hermana del backend).")
        parser.add_argument("--scale", type=float, default=1.0,
                            help="Multiplicador del volumen de ventas sintéticas (def. 1.0).")
        parser.add_argument("--no-fresh", action="store_true",
                            help="No borrar la historia transaccional existente antes de generar.")
        parser.add_argument("--purge-demo", action="store_true",
                            help="Eliminar también los productos/clientes de seed_demo_data.")
        parser.add_argument("--seed", type=int, default=42,
                            help="Semilla aleatoria para reproducibilidad (def. 42).")
        parser.add_argument("--max-customers", type=int, default=0,
                            help="Limitar el nº de clientes importados (0 = todos).")

    def handle(self, *args, **opt):
        # La consola de Windows usa cp1252 y no puede codificar algunos caracteres
        # (p. ej. la flecha →). Forzamos UTF-8 en la salida para evitar UnicodeEncodeError.
        import sys
        for stream in (sys.stdout, sys.stderr):
            try:
                stream.reconfigure(encoding="utf-8")
            except (AttributeError, ValueError):
                pass

        random.seed(opt["seed"])
        self.rng = random.Random(opt["seed"])
        # RNG aislado para el servicio de Mantenimiento y las fechas de alta de clientes:
        # estos se generan al FINAL, así que `self.rng` (la data principal) queda idéntica
        # bit a bit y el servicio solo añade una capa suave que no afecta la exactitud.
        self._svc_rng = random.Random(opt["seed"] + 7)
        self.rates = RateModel()
        openpyxl = _load_openpyxl()

        res = Path(opt["resources"]) if opt["resources"] else (settings.BASE_DIR.parent / "resources")
        if not res.exists():
            raise CommandError(f"No se encontró la carpeta resources/ en: {res}")
        self.stdout.write(self.style.MIGRATE_HEADING(f"Leyendo recursos desde: {res}"))

        with transaction.atomic():
            if not opt["no_fresh"]:
                self._wipe(purge_demo=opt["purge_demo"])

            cats = self._ensure_categories()
            products = self._import_products(openpyxl, res, cats)
            self._import_customers(openpyxl, res, opt["max_customers"])
            sellers = self._ensure_sellers()
            self._build_exchange_rates()

            active = self._active_customer_pool()
            admin = User.objects.filter(is_superuser=True).order_by("id").first()

            # Ventas directas + oportunidades de presupuesto (cuya conversión depende de
            # features y, al convertir, genera su venta). Se persiste todo junto para que
            # el inventario y los FKs presupuesto->venta queden consistentes.
            direct_sales = self._build_direct_sales(active, sellers, scale=opt["scale"])
            quote_specs, converted_sales = self._build_quote_opportunities(active, sellers)
            all_sales = direct_sales + converted_sales
            self._persist_sales(all_sales)
            self._persist_quotes(quote_specs)
            self._build_inventory(products, all_sales, admin)
            self._build_price_history(products)

            # Servicio de Mantenimiento (precio flexible) + sus ventas históricas. Se
            # genera al final, con un RNG propio y sin tocar inventario, de modo que la
            # data principal queda intacta y solo se suma una capa suave a las series.
            self._build_maintenance(active, sellers)

            # Fechas de alta (created_at) verosímiles para distinguir clientes nuevos de
            # antiguos en los paneles (no alimenta ningún modelo de ML).
            self._set_customer_registration_dates()

        self._summary()

    # ---------------------------------------------------------------- wipe -- #
    def _wipe(self, *, purge_demo):
        self.stdout.write("Borrando ventas, productos e historia transaccional previa…")
        InventoryMovement.objects.all().delete()
        QuoteItem.objects.all().delete()
        Quote.objects.all().delete()
        SaleItem.objects.all().delete()
        Sale.objects.all().delete()
        ProductPriceHistory.objects.all().delete()
        # Se borra TODO el catálogo para reemplazarlo por la lista de precios vigente.
        # Las filas de competencia que apuntaban a un producto quedan en NULL
        # (FK SET_NULL); se pueden re-asociar luego con `manage.py rematch_products`.
        Product.objects.all().delete()
        ExchangeRate.objects.all().delete()
        # Reiniciar la condición de cliente activo para que la promoción sea idempotente.
        Customer.objects.update(is_active_customer=False)
        if purge_demo:
            # Productos/clientes de seed_demo_data (SKUs OK-/VIS-/ESC-/ARC-/MES- y RIFs demo).
            Product.objects.filter(sku__regex=r"^(OK-|VIS-|ESC-\d|ARC-|MES-)").delete()
            Customer.objects.filter(rif__in=[
                "J-12345678-9", "J-29876543-1", "G-20000123-4", "V-15678234-0", "J-31122334-5",
            ]).delete()

    # ---------------------------------------------------------- categorías -- #
    def _ensure_categories(self):
        cats = {}
        for name, _prefix in CATEGORIES:
            cat, _ = Category.objects.get_or_create(name=name, defaults={"slug": slugify(name)})
            cats[name] = cat
        self.stdout.write(self.style.SUCCESS(f"Categorías: {len(cats)}"))
        return cats

    # ------------------------------------------------------------ productos -- #
    def _import_products(self, openpyxl, res, cats):
        # norm_name -> (display_name, sale_usd, colors, material_text|None)
        merged = {}
        price_path = res / PRICE_LIST_FILE
        if not price_path.exists():
            raise CommandError(f"No se encontró la lista de precios vigente: {price_path}")
        for name, sell, colors, material in read_new_price_list(openpyxl, price_path):
            merged[norm_name(name)] = (name, sell, colors, material)
        # Sólo los productos de la lista de precios vigente quedan ACTIVOS (es el
        # catálogo actual de la empresa).
        catalog_keys = set(merged)
        # Productos que sólo aparecen en las ventas reales (no en la lista vigente):
        # se conservan para que esas ventas resuelvan a un producto real, pero quedan
        # INACTIVOS (están descontinuados: ya no figuran en el catálogo vigente).
        for name, _buy, sell in CURATED_EXTRA_PRODUCTS:
            if norm_name(name) not in merged:
                merged[norm_name(name)] = (name, float(sell), [], None)

        counters = {prefix: 0 for prefix in CAT_PREFIX.values()}
        created = 0
        self._products = []
        for _key, (name, sell, colors, material_text) in sorted(merged.items()):
            cat_name, mat_default, manufactured, min_stock = classify_product(name)
            material = map_material(material_text) if material_text else mat_default
            # La empresa está cara: el costo se deriva de un margen bruto objetivo
            # (~33% ± jitter), por lo que la utilidad es una fracción estable del ingreso.
            margin = GROSS_MARGIN + self.rng.uniform(-GROSS_MARGIN_JITTER, GROSS_MARGIN_JITTER)
            buy = sell * (1.0 - margin)
            prefix = CAT_PREFIX[cat_name]
            counters[prefix] += 1
            sku = f"MSC-{prefix}-{counters[prefix]:03d}"
            display = title_es(name)
            prod, was_created = Product.objects.update_or_create(
                sku=sku,
                defaults=dict(
                    name=trunc(display, 100), full_name=trunc(display, 255),
                    category=cats[cat_name], material=material,
                    colors=colors or CAT_COLORS.get(cat_name, []),
                    purchase_price_usd=d2(buy), sale_price_usd=d2(sell),
                    min_stock=min_stock, is_manufactured=manufactured,
                    is_active=_key in catalog_keys,
                ),
            )
            created += int(was_created)
            self._products.append(prod)
        self._assign_popularity()
        self.stdout.write(self.style.SUCCESS(
            f"Productos: {created} nuevos (catálogo total {len(self._products)})."))
        return self._products

    # Popularidad relativa por producto: las sillas económicas/operativas se venden mucho
    # más que un módulo de recepción. Da al modelo de demanda una señal por producto.
    _CAT_POPULARITY = {
        "Sillas de Visita": 1.3, "Sillas Operativas y de Cajero": 1.25,
        "Sillas Plásticas": 1.2, "Sillas Ejecutivas": 1.15,
        "Sillas Presidenciales": 0.95, "Escritorios": 1.0, "Mesas": 0.8,
        "Bibliotecas y Archivadores": 0.7, "Recepción y Módulos": 0.55,
        "Accesorios y Repuestos": 0.6,
    }

    def _assign_popularity(self):
        # Popularidad DETERMINISTA (sin azar) para que los productos "estrella" sean
        # estables entre corridas y su demanda mensual sea de mayor magnitud y suave
        # (menos ruido de Poisson) → el modelo de demanda la aprende mejor. El ~20% con
        # mayor puntaje (sillas económicas de categorías populares) recibe un ×4.5.
        base_scores = []
        for p in self._products:
            price = float(p.sale_price_usd)
            tier = 3.0 if price <= 80 else 1.6 if price <= 200 else 0.7
            cf = self._CAT_POPULARITY.get(p.category.name, 1.0)
            base_scores.append(tier * cf)
        order = sorted(range(len(base_scores)), key=lambda i: base_scores[i], reverse=True)
        n_star = max(1, int(round(len(base_scores) * 0.20)))
        stars = set(order[:n_star])
        self._pop_weights = [
            max(0.03, s * (4.5 if i in stars else 1.0)) for i, s in enumerate(base_scores)
        ]

    def _weighted_sample(self, k):
        """Muestra k productos sin reemplazo, ponderando por popularidad."""
        items, weights = list(self._products), list(self._pop_weights)
        out = []
        for _ in range(min(k, len(items))):
            idx = self.rng.choices(range(len(items)), weights=weights)[0]
            out.append(items.pop(idx))
            weights.pop(idx)
        return out

    def _mom_growth(self, y, m):
        """Crecimiento mensual de la tasa paralela (mes anterior -> mes ``m``)."""
        cur = float(self.rates.for_date(date(y, m, 15))[1])
        py, pm = (y - 1, 12) if m == 1 else (y, m - 1)
        prev = float(self.rates.for_date(date(py, pm, 15))[1])
        return (cur / prev - 1.0) if prev else 0.0

    def _rate_shock(self, d):
        """Shock cambiario en [0,1]: cuánto SUPERA la devaluación del mes a su norma reciente.

        Se mide como *desviación* del crecimiento de los 3 meses previos, no como crecimiento
        absoluto: así un mes de inflación estable (p. ej. el 15%/mes sostenido de 2025) marca
        ~0, y sólo los *saltos* bruscos (ago-2022, ene-2023, fin-2024) marcan shock. El efecto
        sobre demanda/conversión queda separado de la tendencia de crecimiento del negocio."""
        cache = self.__dict__.setdefault("_shock_cache", {})
        key = (d.year, d.month)
        if key not in cache:
            g = self._mom_growth(d.year, d.month)
            prevs, yy, mm = [], d.year, d.month
            for _ in range(3):
                mm -= 1
                if mm == 0:
                    mm, yy = 12, yy - 1
                prevs.append(self._mom_growth(yy, mm))
            norm = max(0.04, sum(prevs) / len(prevs))
            cache[key] = max(0.0, min(1.0, (g - norm) / 0.20))
        return cache[key]

    def _affordability(self, d, segment):
        """Multiplicador de demanda por poder de compra (cae con el shock cambiario).

        El detal (discrecional, sensible al precio) cae más que lo institucional
        (proyectos ya contratados). Como el shock se deriva de la tasa, el modelo lo
        ve por su variable ``shock_cambiario`` y la caída queda aprendible."""
        return 1.0 - SHOCK_SENSITIVITY[segment] * self._rate_shock(d)

    @staticmethod
    def _month_index(y, m):
        """Índice de mes desde 2022-01 (0-based)."""
        return (y - 2022) * 12 + (m - 1)

    @staticmethod
    def _interp(points: dict, t: float) -> float:
        """Interpolación lineal entre puntos de control {índice: valor}."""
        keys = sorted(points)
        if t <= keys[0]:
            return points[keys[0]]
        if t >= keys[-1]:
            return points[keys[-1]]
        for a, b in zip(keys, keys[1:]):
            if a <= t <= b:
                return points[a] + (points[b] - points[a]) * (t - a) / (b - a)
        return points[keys[-1]]

    def _resolve_product(self, fuzzy_name):
        """Resuelve un nombre suelto (de una venta real) al producto más parecido."""
        target = norm_name(fuzzy_name)
        toks = set(target.split())
        best, best_score = None, 0.0
        for p in self._products:
            pn = norm_name(p.name)
            ptoks = set(pn.split())
            if not ptoks:
                continue
            inter = len(toks & ptoks)
            if inter == 0 and target not in pn and pn not in target:
                continue
            score = inter / max(1, len(toks | ptoks))
            if target in pn or pn in target:
                score += 0.5
            if score > best_score:
                best, best_score = p, score
        return best or self.rng.choice(self._products)

    # ------------------------------------------------------------ clientes -- #
    def _import_customers(self, openpyxl, res, max_customers):
        path = res / "Base de datos de clientes  y prospectos.xlsx"
        rows = read_customers(openpyxl, path)
        existing = set(Customer.objects.values_list("rif", flat=True))
        seen = set(existing)
        to_create = []
        for r in rows:
            r = list(r) + [None] * (14 - len(r))
            rif = normalize_rif(r[0])
            company = trunc(r[1], 200)
            if not rif or not company or rif in seen:
                continue
            seen.add(rif)
            employees = None
            try:
                employees = int(r[13]) if r[13] not in (None, "") else None
            except (ValueError, TypeError):
                employees = None
            to_create.append(Customer(
                rif=rif, company_name=company,
                customer_type=classify_customer_type(rif, company),
                sector=trunc(r[2], 100),
                contact_first_name=trunc(r[3], 100), contact_last_name=trunc(r[4], 100),
                contact_ci=trunc(r[5], 15), phone=clean_phone(r[6]), mobile=clean_phone(r[7]),
                email=clean_email(r[8]),
                state=trunc(str(r[9]).title() if r[9] else "", 100),
                municipality=trunc(r[10], 100), parish=trunc(r[11], 100),
                fiscal_address=(str(r[12]).strip() if r[12] else ""),
                total_employees=employees, is_active_customer=False,
            ))
            if max_customers and len(to_create) >= max_customers:
                break
        Customer.objects.bulk_create(to_create, batch_size=500, ignore_conflicts=True)
        self.stdout.write(self.style.SUCCESS(
            f"Clientes/prospectos importados: {len(to_create)} (total en BD ahora)."))

        # Compradores reales → clientes activos.
        real_rifs = []
        for i, (name, ctype, state, muni) in enumerate(REAL_BUYERS, start=1):
            rif = REAL_BUYER_RIFS.get(name, f"J-9000{i:04d}-0")
            real_rifs.append(rif)
            Customer.objects.get_or_create(
                rif=rif,
                defaults=dict(company_name=trunc(name, 200), customer_type=ctype,
                              state=state, municipality=muni,
                              sector="Mobiliario / cliente recurrente"),
            )
        # Promover una fracción de prospectos a clientes activos (cartera de la empresa).
        # ``order_by`` garantiza un muestreo reproducible entre ejecuciones.
        prospect_rifs = list(Customer.objects.filter(is_active_customer=False)
                             .exclude(rif__in=real_rifs).order_by("rif")
                             .values_list("rif", flat=True))
        n_active = min(len(prospect_rifs), 420)
        promote = self.rng.sample(prospect_rifs, n_active) if prospect_rifs else []
        Customer.objects.filter(rif__in=promote + real_rifs).update(is_active_customer=True)
        self.stdout.write(self.style.SUCCESS(
            f"Clientes activos: {Customer.objects.filter(is_active_customer=True).count()}."))

    def _active_customer_pool(self):
        pool = list(Customer.objects.filter(is_active_customer=True))
        self.rng.shuffle(pool)
        return pool

    # ---------------------------------------------------------- vendedores -- #
    def _ensure_sellers(self):
        sellers = []
        for fn, ln, email, comm in REAL_SELLERS + EXTRA_SELLERS:
            s, _ = Seller.objects.get_or_create(
                first_name=fn, last_name=ln,
                defaults=dict(email=email, commission_rate=comm, is_active=True),
            )
            sellers.append(s)
        # Ex-vendedores: se crean INACTIVOS y se guardan con su fecha de salida para que
        # sólo reciban ventas/presupuestos mientras estuvieron en la empresa.
        self._former_sellers = []
        for fn, ln, email, comm, left_on in FORMER_SELLERS:
            s, _ = Seller.objects.get_or_create(
                first_name=fn, last_name=ln,
                defaults=dict(email=email, commission_rate=comm, is_active=False),
            )
            self._former_sellers.append((s, left_on))
        # Vendedor ligado al admin (si existe): se crea SOLO para que el admin pueda
        # registrar ventas desde la UI, pero se EXCLUYE de la fuerza de ventas seedeada
        # (no se le atribuye ninguna venta ni presupuesto histórico).
        admin = User.objects.filter(is_superuser=True).order_by("id").first()
        if admin:
            Seller.objects.get_or_create(
                user=admin,
                defaults=dict(first_name=admin.first_name or "Admin",
                              last_name=admin.last_name or "Maescar",
                              email=admin.email or "admin@maescar.com",
                              commission_rate=Decimal("10.00")),
            )
        self._sellers_by_name = {s.first_name: s for s in sellers}
        self.stdout.write(self.style.SUCCESS(
            f"Vendedores: {len(sellers)} activos + {len(self._former_sellers)} ex-vendedores "
            f"(inactivos, con ventas sólo hasta su salida); el admin queda solo para registro manual."))
        return sellers

    # ------------------------------------------------------ tasas de cambio -- #
    def _build_exchange_rates(self):
        objs = []
        for key in sorted(BCV_ANCHORS):
            d = date(int(key[:4]), int(key[5:]), 1)
            bcv, par = self.rates.for_date(d)
            objs.append(ExchangeRate(date=d, bcv_rate=d4(BCV_ANCHORS[key]),
                                     parallel_rate=par, source=ExchangeRate.SourceChoices.BCV))
        # Fila del día de hoy con el dato real actual.
        objs.append(ExchangeRate(date=TODAY_RATE[0], bcv_rate=TODAY_RATE[1],
                                 parallel_rate=TODAY_RATE[2], source=ExchangeRate.SourceChoices.BCV))
        ExchangeRate.objects.bulk_create(objs, ignore_conflicts=True)
        self.stdout.write(self.style.SUCCESS(f"Tasas de cambio: {len(objs)} (2022 a hoy)."))

    # ----------------------------------------------------------- ventas ---- #
    def _make_sale_dict(self, d, customer, seller, items_spec, *, sale_type, status=None):
        """items_spec: lista de (producto, cant, precio_lista, unit_sale, unit_cost, disc_pct)."""
        bcv, par = self.rates.for_date(d)
        items = []
        total_sale = total_cost = total_disc = Decimal("0")
        for prod, qty, list_price, usale, ucost, disc_pct in items_spec:
            list_price, usale, ucost = d2(list_price), d2(usale), d2(ucost)
            sub_s, sub_c = d2(usale * qty), d2(ucost * qty)
            line_disc = max(Decimal("0"), d2((list_price - usale) * qty))
            items.append(dict(product=prod, quantity=qty, list_price=list_price,
                              discount_pct=d2(disc_pct), unit_sale=usale, unit_cost=ucost,
                              sub_sale=sub_s, sub_cost=sub_c, profit=d2(sub_s - sub_c)))
            total_sale += sub_s
            total_cost += sub_c
            total_disc += line_disc
        profit = d2(total_sale - total_cost)
        commission = d2(profit * seller.commission_rate / Decimal("100"))
        if status is None:
            roll = self.rng.random()
            status = (Sale.StatusChoices.CANCELLED if roll < 0.04
                      else Sale.StatusChoices.PENDING if roll < 0.08
                      else Sale.StatusChoices.COMPLETED)
        return dict(customer=customer, seller=seller, sale_date=d, sale_type=sale_type,
                    status=status, total_sale_usd=d2(total_sale), total_cost_usd=d2(total_cost),
                    total_profit_usd=profit, total_discount_usd=d2(total_disc),
                    total_sale_ves=d2(total_sale * par), commission_usd=commission,
                    bcv_rate=bcv, parallel_rate=par, items=items)

    def _basket(self, d, segment):
        """Cesta verosímil según el segmento.

        Detal: pocas líneas y unidades, descuento bajo. Institucional (proyectos):
        más líneas y por lotes, descuento mayor. Devuelve tuplas de 6 elementos
        ``(producto, cant, precio_lista, precio_neto, costo, disc_pct)``."""
        factor = price_factor(d)
        ref = float(PRICE_FACTOR_START)
        disc_mean, disc_sd = DISCOUNT_BANDS[segment]
        if segment == "retail":
            n_lines = self.rng.choices([1, 2], weights=[68, 32])[0]
        else:
            n_lines = self.rng.choices([1, 2, 3], weights=[40, 38, 22])[0]
        chosen = self._weighted_sample(n_lines)        # ponderado por popularidad
        spec = []
        for prod in chosen:
            list_price = Decimal(prod.sale_price_usd) * factor
            base_cost = Decimal(prod.purchase_price_usd or prod.sale_price_usd * Decimal("0.67")) * factor
            disc = min(0.40, max(0.0, self.rng.gauss(disc_mean, disc_sd)))
            usale = list_price * Decimal(str(1 - disc))
            ucost = base_cost * Decimal(str(self.rng.uniform(0.995, 1.005)))  # ruido de costo reducido ±1% -> ±0,5%
            # Cantidades acotadas: ninguna venta domina el total del mes (ingreso suave).
            if segment == "retail":
                if prod.sale_price_usd <= 80:
                    qty = self.rng.choices([1, 2, 3], weights=[52, 32, 16])[0]
                else:
                    qty = self.rng.choices([1, 2], weights=[80, 20])[0]
            else:  # institucional: lotes moderados de proyecto
                if prod.sale_price_usd <= 80:
                    qty = self.rng.choices([3, 4, 6, 8], weights=[26, 30, 26, 18])[0]
                elif prod.sale_price_usd <= 200:
                    qty = self.rng.choices([2, 3, 4, 5], weights=[34, 30, 22, 14])[0]
                else:
                    qty = self.rng.choices([1, 2, 3], weights=[52, 32, 16])[0]
            # Elasticidad precio->cantidad: a mayor precio relativo (vs base 2022), menos unidades.
            rel = (float(factor) * (1 - disc)) / ref
            qty = max(1, int(round(qty * rel ** DEMAND_PRICE_ELASTICITY)))
            spec.append((prod, qty, list_price, usale, ucost, round(disc * 100, 2)))
        return spec

    # Calendario de actividad comercial (compartido por ventas y presupuestos).
    _SEASONAL = {1: 1.15, 2: 1.0, 3: 1.1, 4: 0.95, 5: 1.0, 6: 0.9,
                 7: 0.85, 8: 1.05, 9: 1.15, 10: 1.1, 11: 1.05, 12: 0.9}

    @staticmethod
    def _month_last_day(y, m):
        last = (date(y + (m // 12), (m % 12) + 1, 1) - timedelta(days=1)).day
        if (y, m) == (SYNTH_END.year, SYNTH_END.month):
            last = min(last, SYNTH_END.day)
        return last

    def _segment_target_revenue(self, y, m, segment, scale):
        """Ingreso USD OBJETIVO del mes para un segmento: base × tendencia(segmento) ×
        estacionalidad × poder de compra (shock) × ruido pequeño. La generación rellena
        ventas hasta acercarse a este objetivo → serie de ingreso suave y aprendible."""
        t = self._month_index(y, m)
        if segment == "retail":
            trend = self._interp(RETAIL_TREND, t)
            base = RETAIL_REV_BASE
        else:
            trend = (1.0 + INSTITUTIONAL_GROWTH) ** t   # exponencial: tasa constante
            base = INSTITUTIONAL_REV_BASE
        target = base * trend * self._SEASONAL[m] * scale * self._affordability(date(y, m, 15), segment)
        return max(0.0, self.rng.gauss(target, target * 0.01))   # ruido mensual reducido 0,02 -> 0,01

    def _quote_count(self, y, m, scale):
        """Nº de presupuestos (proyectos) del mes — pipeline del clasificador."""
        t = self._month_index(y, m)
        trend = (1.0 + INSTITUTIONAL_GROWTH) ** t
        n = BASE_QUOTES * trend * self._SEASONAL[m] * scale * self._affordability(date(y, m, 15), "institutional")
        return max(0, int(self.rng.gauss(n, n * 0.12)))

    def _segment_pools(self, active):
        """Separa la cartera activa en detal (particulares + empresas) e institucional
        (instituciones + empresas/proyectos). El solape en empresas es intencional."""
        retail = [c for c in active if c.customer_type != Customer.TypeChoices.INSTITUTIONAL]
        inst = [c for c in active if c.customer_type in (
            Customer.TypeChoices.INSTITUTIONAL, Customer.TypeChoices.CORPORATE)]
        return (retail or active), (inst or active)

    def _seller_weights(self, sellers):
        """Peso al repartir ventas/presupuestos: los vendedores senior llevan más cartera
        que el resto del equipo (≈50% para los dos senior, el resto entre los demás)."""
        return [5 if s.first_name in SENIOR_SELLERS else 2 for s in sellers]

    def _sellers_for_month(self, sellers, y, m):
        """Vendedores que estaban en la empresa en el mes (y, m) con sus pesos: los activos
        siempre, los ex-vendedores sólo hasta su fecha de salida (rotación de personal)."""
        first = date(y, m, 1)
        pool = list(sellers) + [s for (s, left_on) in self._former_sellers if first <= left_on]
        return pool, self._seller_weights(pool)

    def _build_direct_sales(self, active, sellers, *, scale):
        """Ventas directas por segmento (detal flojo, institucional sano): en memoria."""
        self.stdout.write("Generando ventas (detal + institucional) 2022 → abril 2026...")
        sale_dicts = []

        # 1) Ventas reales seedeadas (ene–feb 2022). Sin descuento (precio = lista).
        cust_by_name = {c.company_name: c for c in Customer.objects.filter(is_active_customer=True)}
        for d, buyer, seller_name, items in REAL_SALES:
            customer = cust_by_name.get(buyer) or self.rng.choice(active)
            seller = self._sellers_by_name.get(seller_name, sellers[0])
            spec = [(self._resolve_product(pn), qty, d2(usale), d2(usale), d2(ucost), Decimal("0.00"))
                    for (pn, qty, usale, ucost) in items]
            total = sum((s[3] * s[1] for s in spec), Decimal("0"))
            units = sum(s[1] for s in spec)
            stype = (Sale.TypeChoices.INSTITUTIONAL if (total >= 600 or units >= 6)
                     else Sale.TypeChoices.RETAIL)
            sale_dicts.append(self._make_sale_dict(d, customer, seller, spec,
                                                   sale_type=stype,
                                                   status=Sale.StatusChoices.COMPLETED))

        # 2) Ventas sintéticas mes a mes, rellenando hasta el ingreso objetivo de
        #    cada segmento (ingreso suave, con ventas pequeñas que no dominan el mes).
        retail_pool, inst_pool = self._segment_pools(active)
        streams = (("retail", retail_pool, Sale.TypeChoices.RETAIL),
                   ("institutional", inst_pool, Sale.TypeChoices.INSTITUTIONAL))
        for y, m in iter_months(date(2022, 1, 1), SYNTH_END):
            last_day = self._month_last_day(y, m)
            month_sellers, month_weights = self._sellers_for_month(sellers, y, m)
            for segment, pool, stype in streams:
                target = self._segment_target_revenue(y, m, segment, scale)
                acc = 0.0
                guard = 0
                while acc < target and guard < 400:
                    guard += 1
                    d = date(y, m, self.rng.randint(1, last_day))
                    customer = self.rng.choice(pool)
                    seller = self.rng.choices(month_sellers, weights=month_weights)[0]
                    sd = self._make_sale_dict(d, customer, seller,
                                              self._basket(d, segment), sale_type=stype)
                    sale_dicts.append(sd)
                    acc += float(sd["total_sale_usd"])
        return sale_dicts

    def _persist_sales(self, sale_dicts):
        """Persiste cabeceras y líneas; deja en cada dict su instancia ``obj`` para los FKs."""
        sale_dicts.sort(key=lambda s: s["sale_date"])
        sale_objs = [Sale(**{k: v for k, v in sd.items() if k not in ("items", "obj")})
                     for sd in sale_dicts]
        Sale.objects.bulk_create(sale_objs, batch_size=500)
        item_objs = []
        for sd, sale in zip(sale_dicts, sale_objs):
            sd["obj"] = sale
            for it in sd["items"]:
                item_objs.append(SaleItem(
                    sale=sale, product=it["product"], quantity=it["quantity"],
                    unit_list_price_usd=it["list_price"], discount_pct=it["discount_pct"],
                    unit_sale_price_usd=it["unit_sale"], unit_cost_price_usd=it["unit_cost"],
                    subtotal_sale_usd=it["sub_sale"], subtotal_cost_usd=it["sub_cost"],
                    line_profit_usd=it["profit"]))
        SaleItem.objects.bulk_create(item_objs, batch_size=1000)
        self.stdout.write(self.style.SUCCESS(
            f"Ventas: {len(sale_objs)} (con {len(item_objs)} líneas)."))

    # -------------------------------------------------------- inventario --- #
    def _build_inventory(self, products, sale_dicts, admin):
        self.stdout.write("Generando movimientos de inventario…")
        # Eventos de salida por producto (sólo ventas que mueven stock).
        demand = {p.id: [] for p in products}
        for sd in sale_dicts:
            for it in sd["items"]:
                demand[it["product"].id].append(
                    (sd["sale_date"], it["quantity"], sd["obj"], sd["status"]))

        movements = []
        stock_by_id = {}
        low_stock_ids = set(self.rng.sample([p.id for p in products],
                                            k=max(1, len(products) // 5)))  # ~20% en stock bajo
        prod_by_id = {p.id: p for p in products}

        for pid, events in demand.items():
            prod = prod_by_id[pid]
            events.sort(key=lambda e: e[0])
            total_q = sum(e[1] for e in events) or 0
            if events:
                months_active = max(1, (events[-1][0] - events[0][0]).days / 30.0)
                avg_month = total_q / months_active
            else:
                avg_month = 1.0
            buffer = max(prod.min_stock * 2, int(round(avg_month * 2)) + 1)

            running = 0
            for (d, q, sale, status) in events:
                if running < q:
                    lot = (q - running) + buffer
                    ent_date = d - timedelta(days=self.rng.randint(2, 8))
                    if ent_date < date(2022, 1, 1):
                        ent_date = date(2022, 1, 1)
                    movements.append(InventoryMovement(
                        product=prod, movement_type=InventoryMovement.MovementTypeChoices.ENTRY,
                        quantity=lot, movement_date=ent_date,
                        reference=f"Compra N° {self.rng.randint(1000, 9999)}",
                        responsible=admin, notes=RESTOCK_NOTES))
                    running += lot
                # Salida por venta.
                movements.append(InventoryMovement(
                    product=prod, movement_type=InventoryMovement.MovementTypeChoices.EXIT,
                    quantity=-q, movement_date=d, sale=sale, responsible=admin,
                    notes=f"Salida por venta #{sale.pk}"))
                running -= q
                if status == Sale.StatusChoices.CANCELLED:
                    # Devolución que revierte la salida (como una anulación).
                    movements.append(InventoryMovement(
                        product=prod, movement_type=InventoryMovement.MovementTypeChoices.RETURN,
                        quantity=q, movement_date=d + timedelta(days=self.rng.randint(1, 5)),
                        sale=sale, responsible=admin, notes="Devolución por anulación de venta"))
                    running += q

            # Ajuste final para fijar el stock actual (algunos productos en stock bajo).
            if pid in low_stock_ids:
                target = self.rng.randint(0, max(0, prod.min_stock - 1))
            else:
                target = running  # se queda con el buffer remanente
            if target != running:
                diff = target - running
                movements.append(InventoryMovement(
                    product=prod, movement_type=InventoryMovement.MovementTypeChoices.ADJUSTMENT,
                    quantity=diff, movement_date=SYNTH_END,
                    responsible=admin, notes="Ajuste por conteo físico de inventario"))
                running = target
            stock_by_id[pid] = running

        InventoryMovement.objects.bulk_create(movements, batch_size=1000)
        for p in products:
            p.stock = max(0, stock_by_id.get(p.id, 0))
        Product.objects.bulk_update(products, ["stock"], batch_size=500)
        low = sum(1 for p in products if p.stock <= p.min_stock)
        self.stdout.write(self.style.SUCCESS(
            f"Movimientos de inventario: {len(movements)}. Productos en stock bajo: {low}."))

    # --------------------------------------------------- historial precios -- #
    def _build_price_history(self, products, rng=None):
        rng = rng or self.rng
        rows = []
        points = []
        d = date(2022, 1, 1)
        while d <= SYNTH_END:
            points.append(d)
            # Mensual: en Venezuela los precios se reajustan seguido por la inflación, así
            # la serie de precio en USD es una rampa suave y aprendible (antes ~cada 5 meses,
            # que la dejaba escalonada y de baja varianza → R² limitado).
            month = d.month + 1
            year = d.year + (month - 1) // 12
            month = (month - 1) % 12 + 1
            d = date(year, month, 1)
        for p in products:
            for pd in points:
                # Pequeño ruido mensual (±0,3%) sobre la rampa: la serie de precio queda
                # claramente aprendible pero no perfecta (evita un R² irreal ~0,99).
                f = price_factor(pd) * Decimal(str(rng.uniform(0.997, 1.003)))
                bcv, par = self.rates.for_date(pd)
                buy = d2(Decimal(p.purchase_price_usd or 0) * f)
                sell = d2(Decimal(p.sale_price_usd) * f)
                rows.append(ProductPriceHistory(
                    product=p, purchase_price_usd=buy, sale_price_usd=sell,
                    purchase_price_ves=d2(buy * bcv), sale_price_ves=d2(sell * bcv),
                    bcv_rate=bcv, parallel_rate=par, changed_at=pd,
                    reason=rng.choice(PRICE_HISTORY_REASONS)))
        ProductPriceHistory.objects.bulk_create(rows, batch_size=1000)
        self.stdout.write(self.style.SUCCESS(f"Historial de precios: {len(rows)} registros."))

    # ----------------------------------------------------- mantenimiento --- #
    def _build_maintenance(self, active, sellers):
        """Crea el producto-servicio "Mantenimiento" y su historia de ventas (precio flexible).

        El precio NO es de catálogo: se negocia y se fija en cada venta. El producto solo
        lleva una tarifa de referencia. Las ventas se generan mes a mes con un objetivo
        SUAVE de nº de trabajos (misma tendencia/estacionalidad/shock que el resto, ruido
        bajo) y un precio por trabajo alrededor de una media que sigue la rampa de precios.
        Así la DEMANDA (nº de trabajos = unidades) y el INGRESO/UTILIDAD del servicio son
        series suaves que entran en los modelos sin bajar el R². No mueve inventario.
        """
        rng = self._svc_rng
        cat, _ = Category.objects.get_or_create(
            name=MAINTENANCE_CATEGORY, defaults={"slug": slugify(MAINTENANCE_CATEGORY)})
        ref_cost = MAINTENANCE_REF_PRICE * (1.0 - MAINTENANCE_MARGIN)
        product, _ = Product.objects.update_or_create(
            sku=MAINTENANCE_SKU,
            defaults=dict(
                name=MAINTENANCE_NAME, full_name="Servicio de mantenimiento de mobiliario",
                category=cat, material=Product.MaterialChoices.OTHER, colors=[],
                purchase_price_usd=d2(ref_cost), sale_price_usd=d2(MAINTENANCE_REF_PRICE),
                min_stock=0, stock=0, is_manufactured=False, is_active=True,
            ),
        )

        retail_pool, inst_pool = self._segment_pools(active)
        sale_dicts = []
        for y, m in iter_months(MAINTENANCE_START, SYNTH_END):
            t = self._month_index(y, m)
            last_day = self._month_last_day(y, m)
            trend = (1.0 + MAINTENANCE_GROWTH) ** t                 # crecimiento constante (exponencial)
            afford = self._affordability(date(y, m, 15), "institutional")
            target = MAINTENANCE_BASE_JOBS * trend * self._SEASONAL[m] * afford
            n_jobs = max(0, int(round(rng.gauss(target, target * 0.06))))   # ruido bajo → serie suave
            month_sellers, month_weights = self._sellers_for_month(sellers, y, m)
            mean_price = MAINTENANCE_REF_PRICE * float(price_factor(date(y, m, 15)))
            for _ in range(n_jobs):
                d = date(y, m, rng.randint(1, last_day))
                # El mantenimiento es sobre todo institucional/empresarial (base instalada),
                # con algún particular. El tipo de venta sigue al tipo de cliente.
                if rng.random() < 0.8 and inst_pool:
                    customer = rng.choice(inst_pool)
                else:
                    customer = rng.choice(retail_pool or active)
                seller = rng.choices(month_sellers, weights=month_weights)[0]
                price = max(15.0, rng.gauss(mean_price, mean_price * MAINTENANCE_PRICE_SD))
                margin = MAINTENANCE_MARGIN + rng.uniform(-MAINTENANCE_MARGIN_JITTER, MAINTENANCE_MARGIN_JITTER)
                cost = price * (1.0 - margin)
                stype = (Sale.TypeChoices.INSTITUTIONAL
                         if customer.customer_type != Customer.TypeChoices.INDIVIDUAL
                         else Sale.TypeChoices.RETAIL)
                # spec de 6-tuplas (producto, cant, precio_lista, precio_neto, costo, disc_pct):
                # precio flexible → lista = neto (sin descuento).
                spec = [(product, 1, d2(price), d2(price), d2(cost), Decimal("0.00"))]
                sale_dicts.append(self._make_sale_dict(
                    d, customer, seller, spec, sale_type=stype,
                    status=Sale.StatusChoices.COMPLETED))

        self._persist_sales(sale_dicts)
        # Historial de precios de la tarifa de referencia (rampa suave) con el RNG aislado,
        # para que el modelo de precio también pueda pronosticar el servicio de forma aprendible.
        self._build_price_history([product], rng=self._svc_rng)
        self.stdout.write(self.style.SUCCESS(
            f"Mantenimiento (servicio): {len(sale_dicts)} ventas históricas (precio flexible)."))

    # ----------------------------------------------- fechas de alta clientes -- #
    def _set_customer_registration_dates(self):
        """Fija `created_at` de los clientes para distinguir ALTAS nuevas de antiguas.

        `created_at` es auto_now_add (de fábrica, todos = hoy), así que el panel no podía
        contar "clientes nuevos" por mes. Aquí se reconstruye una fecha de alta verosímil:
        los compradores se dieron de alta poco antes de su PRIMERA compra; los prospectos
        (sin compras) se reparten en el tiempo con sesgo a fechas recientes (la empresa
        capta más leads con los años). No alimenta ningún modelo de ML; es solo para los
        paneles. Determinista (RNG aislado) e idempotente.
        """
        from django.db.models import Min

        rng = self._svc_rng
        first_purchase = {
            r["customer_id"]: r["first"]
            for r in Sale.objects.values("customer_id").annotate(first=Min("sale_date"))
        }
        start = date(2022, 1, 1)
        span_days = max(1, (SYNTH_END - start).days)
        tz = timezone.get_current_timezone()
        updates = []
        for c in Customer.objects.all().only("id"):
            fp = first_purchase.get(c.id)
            if fp:
                reg = max(start, fp - timedelta(days=rng.randint(3, 45)))
            else:
                # Prospecto sin compras: sesgo a fechas recientes (triangular, moda al final).
                reg = start + timedelta(days=int(rng.triangular(0.0, 1.0, 1.0) * span_days))
            dt = datetime.combine(reg, time(9, 0))
            c.created_at = timezone.make_aware(dt, tz) if settings.USE_TZ else dt
            updates.append(c)
        Customer.objects.bulk_update(updates, ["created_at"], batch_size=500)
        self.stdout.write(self.style.SUCCESS(
            f"Fechas de alta de clientes fijadas: {len(updates)}."))

    # ------------------------------------------------------- presupuestos -- #
    def _build_quote_opportunities(self, active, sellers):
        """Genera oportunidades de presupuesto cuya conversión depende de *features*.

        Devuelve ``(quote_specs, converted_sales)``: cada presupuesto convertido genera su
        propia venta (que se añade a la lista de ventas), de modo que el FK presupuesto->venta
        y el inventario quedan consistentes. Así el árbol de decisión tiene una etiqueta
        (convertido / no) con señal real (instalación, monto, tipo de cliente, shock cambiario)."""
        self.stdout.write("Generando presupuestos (proyectos) y conversiones…")
        _retail_pool, inst_pool = self._segment_pools(active)
        specs, converted_sales = [], []
        # El pipeline de presupuestos es la fuente del CLASIFICADOR de conversión; el
        # ingreso institucional ya lo aporta el flujo directo (suave), así que un
        # presupuesto convertido NO genera otra venta (sólo su etiqueta). Sólo los
        # presupuestos recientes (últimos ~55 días) siguen ABIERTOS; los más viejos sin
        # convertir se dan por perdidos (rechazados) — como en una PYME real.
        open_cutoff = SYNTH_END - timedelta(days=55)

        for y, m in iter_months(date(2022, 1, 1), SYNTH_END):
            n = self._quote_count(y, m, 1.0)
            last_day = self._month_last_day(y, m)
            shock = self._rate_shock(date(y, m, 15))
            month_sellers, month_weights = self._sellers_for_month(sellers, y, m)
            for _ in range(n):
                issued = date(y, m, self.rng.randint(1, last_day))
                customer = self.rng.choice(inst_pool)
                seller = self.rng.choices(month_sellers, weights=month_weights)[0]
                basket = self._basket(issued, "institutional")  # 6-tuplas con descuento
                # El presupuesto se cotiza a precio de lista.
                items = [(p, q, d2(list_price)) for (p, q, list_price, _us, _uc, _dp) in basket]
                total_usd = sum((up * q for (_p, q, up) in items), Decimal("0"))
                units = sum(q for (_p, q, _up) in items)
                installation = self.rng.random() < 0.40
                delivery = self.rng.random() < 0.55
                p_conv = conversion_probability(
                    total_usd, units, installation, delivery, customer.customer_type,
                    shock, seller.first_name in SENIOR_SELLERS)
                # Decisión de cierre "afilada" (probit): casi determinista en las
                # features, con algo de ruido. Da una etiqueta separable → el árbol
                # clasificador alcanza buena exactitud (en vez de puro azar de Bernoulli).
                if (p_conv + self.rng.gauss(0.0, 0.04)) > 0.5:   # ruido de etiqueta reducido 0,12 -> 0,04
                    status = Quote.StatusChoices.CONVERTED
                elif issued >= open_cutoff and self.rng.random() < 0.6:
                    # Reciente y sin convertir: una parte sigue en gestión (abierto). El
                    # resto se cierra como rechazado para que el conjunto resuelto mantenga
                    # ambas clases en el tiempo (holdout estable del clasificador).
                    status = self.rng.choices(
                        [Quote.StatusChoices.SENT, Quote.StatusChoices.APPROVED,
                         Quote.StatusChoices.DRAFT], [55, 30, 15])[0]
                else:
                    # Antiguo, o reciente no retenido: se da por perdido (rechazado).
                    status = Quote.StatusChoices.REJECTED
                specs.append(dict(customer=customer, seller=seller, issued=issued, items=items,
                                  installation=installation, delivery=delivery, status=status,
                                  sale_ref=None))

        # El presupuesto real de mayo 2026 (con su tasa documentada).
        rq = REAL_QUOTE
        cust = Customer.objects.filter(company_name=rq["buyer"]).first() or self.rng.choice(active)
        seller = self._sellers_by_name.get("Mariangel", sellers[0])
        items = [(self._resolve_product(pn), qty, d2(price)) for (pn, qty, price) in rq["items"]]
        specs.append(dict(customer=cust, seller=seller, issued=rq["issued"], items=items,
                          installation=rq["installation"], delivery=rq["delivery"],
                          status=Quote.StatusChoices.SENT, sale_ref=None,
                          number=rq["number"], bcv_override=rq["bcv"]))
        return specs, converted_sales

    def _persist_quotes(self, specs):
        specs.sort(key=lambda s: s["issued"])
        quote_objs, metas = [], []
        for i, sp in enumerate(specs, start=1):
            issued = sp["issued"]
            bcv, par = self.rates.for_date(issued)
            if sp.get("bcv_override") is not None:               # tasa documentada (presupuesto real)
                bcv = d4(sp["bcv_override"])
                par = d4(float(sp["bcv_override"]) * RateModel.premium(issued))
            subtotal = sum((d2(up * q) for (_p, q, up) in sp["items"]), Decimal("0"))
            iva = d2(subtotal * Decimal("0.16"))
            total = d2(subtotal + iva)
            number = sp.get("number") or f"{issued.strftime('%d%m%Y')}-{i}"
            conv_sale = sp["sale_ref"]["obj"] if sp.get("sale_ref") else None
            quote_objs.append(Quote(
                quote_number=number, customer=sp["customer"], seller=sp["seller"],
                issued_date=issued, expiry_date=issued + timedelta(days=self.rng.randint(5, 15)),
                bcv_rate=bcv, parallel_rate=par,
                includes_installation=sp["installation"], includes_delivery=sp["delivery"],
                subtotal_usd=d2(subtotal), subtotal_ves=d2(subtotal * bcv),
                iva_rate=Decimal("16.00"), iva_amount_usd=iva,
                total_usd=total, total_ves=d2(total * bcv),
                status=sp["status"], converted_to_sale=conv_sale))
            metas.append(sp["items"])
        Quote.objects.bulk_create(quote_objs, batch_size=500)
        qitems = []
        for q, items in zip(quote_objs, metas):
            for (prod, qty, up) in items:
                qitems.append(QuoteItem(
                    quote=q, product=prod, quantity=qty, unit_price_usd=up,
                    unit_price_ves=d2(up * q.bcv_rate), line_total_usd=d2(up * qty),
                    line_total_ves=d2(up * qty * q.bcv_rate)))
        QuoteItem.objects.bulk_create(qitems, batch_size=1000)
        conv = sum(1 for s in specs if s["status"] == Quote.StatusChoices.CONVERTED)
        self.stdout.write(self.style.SUCCESS(
            f"Presupuestos: {len(quote_objs)} ({conv} convertidos, {len(qitems)} líneas)."))

    # --------------------------------------------------------------- resumen -- #
    def _summary(self):
        from django.db.models import Sum
        self.stdout.write(self.style.MIGRATE_HEADING("\n=== Resumen de la base de datos ==="))
        open_statuses = [Quote.StatusChoices.DRAFT, Quote.StatusChoices.SENT, Quote.StatusChoices.APPROVED]
        retail = Sale.objects.filter(sale_type=Sale.TypeChoices.RETAIL).count()
        inst = Sale.objects.filter(sale_type=Sale.TypeChoices.INSTITUTIONAL).count()
        discount = Sale.objects.aggregate(t=Sum("total_discount_usd"))["t"] or 0
        lines = [
            ("Productos", Product.objects.count()),
            ("Clientes (total)", Customer.objects.count()),
            ("  · activos", Customer.objects.filter(is_active_customer=True).count()),
            ("Vendedores", Seller.objects.count()),
            ("Tasas de cambio", ExchangeRate.objects.count()),
            ("Ventas", Sale.objects.count()),
            ("  · detal", retail),
            ("  · institucional", inst),
            ("Líneas de venta", SaleItem.objects.count()),
            ("Movimientos de inventario", InventoryMovement.objects.count()),
            ("Historial de precios", ProductPriceHistory.objects.count()),
            ("Presupuestos", Quote.objects.count()),
            ("  · abiertos (en gestión)", Quote.objects.filter(status__in=open_statuses).count()),
            ("  · convertidos", Quote.objects.filter(status=Quote.StatusChoices.CONVERTED).count()),
        ]
        for label, n in lines:
            self.stdout.write(f"  {label:.<32} {n:>8,}")
        self.stdout.write(f"  {'Descuento total otorgado (USD)':.<32} {float(discount):>10,.2f}")
        last = ExchangeRate.objects.order_by("-date").first()
        if last:
            self.stdout.write(
                f"\n  Tasa más reciente {last.date}: BCV {last.bcv_rate} | Paralela {last.parallel_rate}")
        self.stdout.write(self.style.SUCCESS("\nBase de datos de Maescar cargada correctamente."))
